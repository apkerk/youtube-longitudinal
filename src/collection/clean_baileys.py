"""
clean_baileys.py
----------------
Data preparation: Clean Bailey's XLSX and produce canonical gender gap panel files.

Input:  data/raw/BAILEYS FULL CHANNEL LIST WITH GENDER AND RACE.xlsx
Output: data/processed/gender_gap_panel_clean.csv        (full cleaned dataset)
        data/channels/gender_gap/channel_ids.csv         (channel_id column only)
        data/channels/gender_gap/channel_metadata.csv    (channel_id + key attributes)

Parsing uses openpyxl with header-based column lookup (NOT positional index).
This avoids the sparse-row bug where XML-based positional parsing skips empty cells.

Usage:
    python -m src.collection.clean_baileys [--test] [--limit N]

Author: Katie Apker
Last Updated: Feb 16, 2026
"""

import argparse
import csv
import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional

# Add parent to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))
import config

# Output directories (derived from config base paths)
PROCESSED_DIR = config.DATA_DIR / "processed"
GENDER_GAP_DIR = config.CHANNELS_DIR / "gender_gap"

# Input file
INPUT_FILE = config.DATA_DIR / "raw" / "BAILEYS FULL CHANNEL LIST WITH GENDER AND RACE.xlsx"

# Race typo corrections
RACE_TYPO_MAP: Dict[str, str] = {
    "blakc": "black",
    "undetermiend": "undetermined",
    "whitee": "white",
    "wht": "white",
    "asain": "asian",
    "wihte": "white",
}

# Metadata columns to extract for the lean channel_metadata.csv
METADATA_COLUMNS: List[str] = [
    "channel_id",
    "perceivedGender",
    "race",
    "runBy",
    "subscriberCount",
    "viewCount",
]

# Expected row count
EXPECTED_CHANNEL_COUNT = 14_169


def setup_logging(test_mode: bool = False) -> logging.Logger:
    """Configure logging with file and stream handlers."""
    config.ensure_directories()
    log_file = config.LOGS_DIR / f"clean_baileys_{config.get_date_stamp()}.log"

    logger = logging.getLogger("clean_baileys")
    logger.setLevel(logging.DEBUG if test_mode else logging.INFO)

    stream_handler = logging.StreamHandler()
    stream_handler.setLevel(logging.INFO)
    stream_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    file_handler = logging.FileHandler(log_file)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))

    logger.addHandler(stream_handler)
    logger.addHandler(file_handler)
    return logger


def clean_string(value: Optional[str]) -> Optional[str]:
    """Strip trailing whitespace from string fields. Pass through None."""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip()
    return value


def clean_race(value: Optional[str]) -> Optional[str]:
    """Normalize race field: strip whitespace, fix known typos, lowercase for matching."""
    if value is None:
        return None
    cleaned = str(value).strip()
    if not cleaned:
        return None
    lookup = cleaned.lower()
    if lookup in RACE_TYPO_MAP:
        return RACE_TYPO_MAP[lookup]
    return cleaned


def load_and_clean(
    logger: logging.Logger,
    limit: Optional[int] = None,
) -> List[Dict]:
    """
    Load Bailey's XLSX and return cleaned row dicts.

    Uses openpyxl with header-based column lookup to avoid
    the sparse-row positional parsing bug.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    logger.info("Loading workbook: %s", INPUT_FILE)
    wb = load_workbook(str(INPUT_FILE), read_only=True, data_only=True)
    ws = wb.active

    # Read header row to build column name -> index mapping
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_names: List[str] = [str(c).strip() if c is not None else f"_unnamed_{i}" for i, c in enumerate(header_row)]
    logger.info("Found %d columns: %s", len(col_names), col_names[:10])

    # Build name -> position map
    col_index: Dict[str, int] = {name: idx for idx, name in enumerate(col_names)}

    rows: List[Dict] = []
    row_count = 0
    for row_tuple in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        if limit is not None and row_count > limit:
            break

        # Build dict by header name
        row_dict: Dict = {}
        for col_name, idx in col_index.items():
            if idx < len(row_tuple):
                raw_val = row_tuple[idx]
            else:
                raw_val = None
            row_dict[col_name] = clean_string(raw_val) if isinstance(raw_val, str) else raw_val

        # Apply race-specific cleaning
        if "race" in row_dict:
            row_dict["race"] = clean_race(row_dict.get("race"))

        rows.append(row_dict)

    wb.close()
    logger.info("Read %d data rows from workbook", len(rows))

    # Deduplicate by channel_id (keep first occurrence)
    seen_ids: set = set()
    unique_rows: List[Dict] = []
    dupes = 0
    for row in rows:
        cid = row.get("channel_id")
        if cid is not None and str(cid).strip() != "":
            cid_str = str(cid).strip()
            if cid_str in seen_ids:
                dupes += 1
                continue
            seen_ids.add(cid_str)
        unique_rows.append(row)

    if dupes > 0:
        logger.info("Removed %d duplicate channel_id rows (%d -> %d)", dupes, len(rows), len(unique_rows))

    return unique_rows


def validate(rows: List[Dict], logger: logging.Logger) -> bool:
    """Validate cleaned data. Returns True if all checks pass."""
    passed = True

    # Check for null channel_ids
    channel_ids = [r.get("channel_id") for r in rows]
    null_ids = sum(1 for cid in channel_ids if cid is None or str(cid).strip() == "")
    if null_ids > 0:
        logger.error("VALIDATION FAILED: %d rows have null/empty channel_id", null_ids)
        passed = False
    else:
        logger.info("PASS: No null channel_ids")

    # Check uniqueness
    non_null_ids = [cid for cid in channel_ids if cid is not None and str(cid).strip() != ""]
    unique_ids = set(non_null_ids)
    duplicates = len(non_null_ids) - len(unique_ids)
    if duplicates > 0:
        logger.error("VALIDATION FAILED: %d duplicate channel_ids found", duplicates)
        passed = False
    else:
        logger.info("PASS: All channel_ids are unique (%d)", len(unique_ids))

    # Check row count (only if not in limit mode)
    if len(rows) != EXPECTED_CHANNEL_COUNT:
        logger.warning(
            "Row count mismatch: got %d, expected %d (may be OK if --limit used)",
            len(rows),
            EXPECTED_CHANNEL_COUNT,
        )
    else:
        logger.info("PASS: Row count matches expected (%d)", EXPECTED_CHANNEL_COUNT)

    return passed


def write_full_clean(rows: List[Dict], logger: logging.Logger) -> Path:
    """Write the full cleaned dataset to CSV."""
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    output_path = PROCESSED_DIR / "gender_gap_panel_clean.csv"

    if not rows:
        logger.error("No rows to write")
        return output_path

    fieldnames = list(rows[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    logger.info("Wrote full clean dataset: %s (%d rows, %d columns)", output_path, len(rows), len(fieldnames))
    return output_path


def write_channel_ids(rows: List[Dict], logger: logging.Logger) -> Path:
    """Write channel_id-only file for panel collection."""
    GENDER_GAP_DIR.mkdir(parents=True, exist_ok=True)
    output_path = GENDER_GAP_DIR / "channel_ids.csv"

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["channel_id"])
        for row in rows:
            cid = row.get("channel_id")
            if cid is not None and str(cid).strip() != "":
                writer.writerow([cid])

    logger.info("Wrote channel IDs: %s (%d rows)", output_path, len(rows))
    return output_path


def write_channel_metadata(rows: List[Dict], logger: logging.Logger) -> Path:
    """Write lean metadata file with key attributes for panel use."""
    GENDER_GAP_DIR.mkdir(parents=True, exist_ok=True)
    output_path = GENDER_GAP_DIR / "channel_metadata.csv"

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=METADATA_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            # Only write columns that exist in METADATA_COLUMNS
            meta_row = {col: row.get(col) for col in METADATA_COLUMNS}
            writer.writerow(meta_row)

    logger.info("Wrote channel metadata: %s (%d rows, %d columns)", output_path, len(rows), len(METADATA_COLUMNS))
    return output_path


def main() -> None:
    """Main entry point for Bailey's data cleaning pipeline."""
    parser = argparse.ArgumentParser(
        description="Clean Bailey's XLSX and produce gender gap panel files."
    )
    parser.add_argument(
        "--test",
        action="store_true",
        help="Run in test mode (verbose logging)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Limit number of rows to process (for testing)",
    )
    args = parser.parse_args()

    logger = setup_logging(test_mode=args.test)

    if args.test:
        logger.info("TEST MODE enabled")
    if args.limit is not None:
        logger.info("Row limit: %d", args.limit)

    # Verify input exists
    if not INPUT_FILE.exists():
        logger.error("Input file not found: %s", INPUT_FILE)
        sys.exit(1)

    # Load and clean
    rows = load_and_clean(logger, limit=args.limit)

    if not rows:
        logger.error("No data loaded. Exiting.")
        sys.exit(1)

    # Validate
    valid = validate(rows, logger)
    if not valid and args.limit is None:
        logger.error("Validation failed on full dataset. Aborting writes.")
        sys.exit(1)

    # Write all three outputs
    full_path = write_full_clean(rows, logger)
    ids_path = write_channel_ids(rows, logger)
    meta_path = write_channel_metadata(rows, logger)

    logger.info("--- Summary ---")
    logger.info("Full clean:       %s", full_path)
    logger.info("Channel IDs:      %s", ids_path)
    logger.info("Channel metadata: %s", meta_path)
    logger.info("Total rows:       %d", len(rows))
    logger.info("Done.")


if __name__ == "__main__":
    main()
